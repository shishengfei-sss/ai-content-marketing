"""自动化测试编排器。

读取《AI内容营销系统-测试用例.xlsx》作为唯一真源，执行已实现的后端 API 用例，
并将结果写回到一份副本 Excel（不破坏原件）：
  - 测试结果（通过 / 失败 / 未实现）
  - 实际结果（接口返回摘要）
  - 测试人 / 测试日期
  - 备注（含规格/实现偏差说明）

运行：
  cd qa-automation
  ..\\..\\  (用 managed python)
  python run.py
"""
import datetime
import json
import os
import traceback

import openpyxl

from config import EXCEL_PATH, RESULT_PATH, TESTER
from cases_api import REGISTRY as REGISTRY_API
from cases_ui import REGISTRY_UI
from ui_helpers import close_browser

REGISTRY = {**REGISTRY_API, **REGISTRY_UI}

# Excel 列（1-based）：L 测试结果 / M 实际结果 / N 测试人 / O 测试日期 / P 备注
COL = {"result": 12, "actual": 13, "tester": 14, "date": 15, "note": 16}

TODAY = datetime.date.today().isoformat()


def load_cases(path):
    wb = openpyxl.load_workbook(path)
    cases = []
    loc = {}  # case_id -> [(ws_title, row)]
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            continue
        header = rows[0]
        if header is None or "用例编号" not in header:
            continue
        cid_col = list(header).index("用例编号")
        for i, r in enumerate(rows[1:], start=2):
            if not r:
                continue
            cid = r[cid_col] if cid_col < len(r) else None
            if cid in (None, ""):
                continue
            cid = str(cid).strip()
            cases.append({
                "sheet": ws.title, "row": i, "case_id": cid,
                "module": r[2] if len(r) > 2 else "",
                "title": r[5] if len(r) > 5 else "",
                "priority": r[9] if len(r) > 9 else "",
                "type": r[10] if len(r) > 10 else "",
                "expected": r[8] if len(r) > 8 else "",
            })
            loc.setdefault(cid, []).append((ws.title, i))
    return wb, cases, loc


def run_all(cases):
    results = {}
    for case in cases:
        cid = case["case_id"]
        if cid in results:
            continue  # 总表与模块表为同一用例，只跑一次
        if cid in REGISTRY:
            try:
                ok, actual, note = REGISTRY[cid]()
                status = "通过" if ok else "失败"
            except Exception as e:  # noqa
                status, actual, note = "失败", f"执行异常: {e}", traceback.format_exc()[:400]
        else:
            if case["type"] == "接口":
                status, actual, note = "未实现", "接口用例，执行器待补充", "需在 cases_api.py 补充实现"
            else:
                status, actual, note = "未实现", "需 UI 自动化(Playwright)", "本框架当前仅覆盖后端 API；UI 用例为下一阶段"
        results[cid] = {"status": status, "actual": str(actual)[:500],
                        "note": str(note)[:500], "case": case}
    return results


def write_back(wb, loc, results):
    counts = {"通过": 0, "失败": 0, "未实现": 0}
    for cid, res in results.items():
        status = res["status"]
        counts[status] = counts.get(status, 0) + 1
        actual = res["actual"]
        note = res["note"]
        for ws_title, row in loc.get(cid, []):
            ws = wb[ws_title]
            ws.cell(row=row, column=COL["result"]).value = status
            ws.cell(row=row, column=COL["actual"]).value = actual
            ws.cell(row=row, column=COL["tester"]).value = TESTER
            ws.cell(row=row, column=COL["date"]).value = TODAY
            ws.cell(row=row, column=COL["note"]).value = note
    return counts


def main():
    print(f"[*] 读取用例真源：{EXCEL_PATH}")
    wb, cases, loc = load_cases(EXCEL_PATH)
    print(f"[*] 解析到 {len(cases)} 条用例（去重后 {len(loc)} 个唯一用例编号）")
    print(f"[*] 已实现执行器：{len(REGISTRY)} 个（{', '.join(sorted(REGISTRY))}）")
    print("[*] 开始执行已实现用例（后端 API + UI）...")
    results = run_all(cases)
    print("[*] 写回结果到副本：", RESULT_PATH)
    counts = write_back(wb, loc, results)
    os.makedirs(os.path.dirname(RESULT_PATH), exist_ok=True)
    wb.save(RESULT_PATH)

    # 控制台摘要
    print("\n================ 执行摘要 ================")
    print(f"  总用例(去重): {len(results)}")
    for k in ("通过", "失败", "未实现"):
        print(f"  {k}: {counts.get(k, 0)}")

    print("\n--- 失败用例 ---")
    for cid, res in results.items():
        if res["status"] == "失败":
            print(f"  [{cid}] {res['case']['title']}")
            print(f"       实际: {res['actual']}")
            if res["note"]:
                print(f"       备注: {res['note']}")

    print("\n--- 规格/实现偏差（已记录到备注）---")
    for cid, res in results.items():
        if "偏差" in res["note"] or "差异" in res["note"] or "修订" in res["note"]:
            print(f"  [{cid}] {res['case']['title']}: {res['note']}")

    # JSON 报告
    report = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": EXCEL_PATH,
        "result_file": RESULT_PATH,
        "summary": counts,
        "total_unique": len(results),
        "cases": [
            {"case_id": cid, **{k: v for k, v in res.items() if k != "case"},
             "module": res["case"]["module"], "title": res["case"]["title"],
             "type": res["case"]["type"], "priority": res["case"]["priority"]}
            for cid, res in results.items()
        ],
    }
    report_path = os.path.join(os.path.dirname(RESULT_PATH), "report.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n[*] JSON 报告：{report_path}")

    # 关闭 Playwright 浏览器（若 UI 用例启动过）
    close_browser()


if __name__ == "__main__":
    main()
