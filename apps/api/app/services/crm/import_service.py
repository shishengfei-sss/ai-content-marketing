"""CRM 数据导入服务。"""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from pathlib import Path
from uuid import UUID

from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.config import settings
from app.dependencies import TenantContext
from app.models.crm import (
    IMPORT_ROW_STATUSES,
    CrmImportJob,
    CrmImportRow,
    Customer,
    Lead,
    Product,
    ProductCategory,
    ProductSpecModel,
    ProductUnit,
    SalesTerritory,
)
from app.schemas.crm import validate_customer_status, validate_lead_status, validate_lead_mobile_value
from app.schemas.crm_deals import (
    ProductCreate,
    ProductSpecModelCreate,
    ProductSpecModelUpdate,
    ProductUpdate,
)
from app.permissions import SYSTEM_ROLE_ADMIN
from app.services.crm.assignment_service import apply_assignment_rules
from app.services.crm.crm_scope_service import assert_can_view_customer, assert_can_view_lead
from app.services.crm.number_service import generate_number
from app.services.crm.product_service import create_product, update_product
from app.services.crm.product_spec_model_service import (
    create_spec_model,
    resolve_spec_model_id_by_name,
    update_spec_model,
)
from app.services.crm.sales_org_service import apply_creator_org_defaults, get_territory
from app.services.crm.schema_service import ensure_entity_schema, list_active_fields, validate_extra_data
from app.services.membership_service import get_membership

IMPORT_ENTITY_TYPES = frozenset({"lead", "customer", "product", "product_spec_model"})

# CSV/Excel 表头常见别名 → field_key（label 变更后仍兼容旧模板）
IMPORT_COLUMN_ALIASES: dict[str, str] = {
    "联系人": "contact_name",
    "规格": "spec_model_id",
    "型号": "spec_model_id",
    "默认税率": "default_tax_rate",
    "默认税率%": "default_tax_rate",
    "税率": "default_tax_rate",
    "税率%": "default_tax_rate",
    "增值税率": "default_tax_rate",
    "标价含税": "price_includes_tax",
    "含税": "price_includes_tax",
    "价格含税": "price_includes_tax",
    "是否含税": "price_includes_tax",
    "销售区域": "territory_id",
    "归属地区": "territory_id",
    "地区": "territory_id",
}
MAX_IMPORT_ROWS = 5000
_PHONE_RE = re.compile(r"^1[3-9]\d{9}$")

LEAD_DB_KEYS = frozenset(
    {
        "company_name",
        "contact_name",
        "mobile",
        "phone",
        "email",
        "source",
        "status",
        "remark",
        "territory_id",
    }
)
CUSTOMER_DB_KEYS = frozenset({"company_name", "mobile", "phone", "email", "status", "remark"})
PRODUCT_DB_KEYS = frozenset(
    {
        "code",
        "name",
        "unit",
        "list_price",
        "cost_price",
        "default_tax_rate",
        "price_includes_tax",
        "category_id",
        "spec_model_id",
        "is_active",
        "description",
    }
)
# 导入时按名称解析的引用字段（模板填中文名）
PRODUCT_REF_NAME_KEYS = frozenset({"category_id", "spec_model_id"})

# 导入模板必填（与前端 ENTITY_REQUIRED_KEYS / 导入校验一致，覆盖存量 schema）
LEAD_IMPORT_REQUIRED_KEYS = frozenset({"company_name", "contact_name", "mobile", "status"})
CUSTOMER_IMPORT_REQUIRED_KEYS = frozenset({"company_name"})
PRODUCT_IMPORT_REQUIRED_KEYS = frozenset({"name"})

SPEC_MODEL_DB_KEYS = frozenset({"name", "code", "description", "sort_order", "is_active"})
SPEC_MODEL_IMPORT_FIELDS: list[tuple[str, str, bool]] = [
    ("name", "名称", True),
    ("code", "编码", False),
    ("description", "说明", False),
    ("sort_order", "排序", False),
    ("is_active", "启用", False),
]
SPEC_MODEL_IMPORT_REQUIRED_KEYS = frozenset({"name"})
SPEC_MODEL_COLUMN_ALIASES: dict[str, str] = {
    "规格型号": "name",
    "规格型号名称": "name",
}


def _import_field_required(entity_type: str, field_key: str, *, schema_required: bool) -> bool:
    # 产品编码导入时不要求：空则 create_product 自动流水号
    if entity_type == "product" and field_key == "code":
        return False
    # 线索销售区域：空则回填创建人主地区，模板不强制 *
    if entity_type == "lead" and field_key == "territory_id":
        return False
    if schema_required:
        return True
    if entity_type == "lead":
        keys = LEAD_IMPORT_REQUIRED_KEYS
    elif entity_type == "customer":
        keys = CUSTOMER_IMPORT_REQUIRED_KEYS
    elif entity_type == "product":
        keys = PRODUCT_IMPORT_REQUIRED_KEYS
    elif entity_type == "product_spec_model":
        keys = SPEC_MODEL_IMPORT_REQUIRED_KEYS
    else:
        keys = frozenset()
    return field_key in keys


def _import_dir(tenant_id: UUID, job_id: UUID) -> Path:
    d = Path(settings.STORAGE_DIR) / "crm_imports" / str(tenant_id) / str(job_id)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _default_options(entity_type: str) -> dict:
    if entity_type == "customer":
        return {
            "duplicate_key": "mobile",
            "on_duplicate": "skip",
            "default_status": "潜在",
            "default_source": "导入",
        }
    if entity_type == "product":
        return {
            "duplicate_key": "code",
            "on_duplicate": "skip",
        }
    if entity_type == "product_spec_model":
        return {
            "duplicate_key": "name",
            "on_duplicate": "skip",
        }
    return {
        "duplicate_key": "mobile",
        "on_duplicate": "skip",
        "default_status": "待跟进",
        "default_source": "导入",
    }


def _db_keys_for(entity_type: str) -> frozenset[str]:
    if entity_type == "lead":
        return LEAD_DB_KEYS
    if entity_type == "customer":
        return CUSTOMER_DB_KEYS
    if entity_type == "product":
        return PRODUCT_DB_KEYS
    if entity_type == "product_spec_model":
        return SPEC_MODEL_DB_KEYS
    return frozenset()


def _resolve_product_category_id(db: Session, tenant_id: UUID, name: str) -> UUID:
    row = (
        db.query(ProductCategory)
        .filter(
            ProductCategory.tenant_id == tenant_id,
            ProductCategory.name == name.strip(),
            ProductCategory.is_active.is_(True),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=400, detail=f"分类「{name}」不存在或已停用")
    return row.id


def _resolve_product_unit_name(db: Session, tenant_id: UUID, name: str) -> str:
    unit_name = name.strip()
    row = (
        db.query(ProductUnit)
        .filter(
            ProductUnit.tenant_id == tenant_id,
            ProductUnit.name == unit_name,
            ProductUnit.is_active.is_(True),
        )
        .first()
    )
    if not row:
        raise HTTPException(status_code=400, detail=f"单位「{name}」不存在或已停用")
    return row.name


def _resolve_territory_id(db: Session, tenant_id: UUID, value: str) -> UUID:
    """按地区名称或 UUID 解析销售区域。"""
    text = str(value).strip()
    if not text:
        raise HTTPException(status_code=400, detail="销售区域不能为空")
    try:
        tid = UUID(text)
    except (TypeError, ValueError):
        tid = None
    if tid is not None:
        row = get_territory(db, tenant_id, tid)
        if not row:
            raise HTTPException(status_code=400, detail=f"销售区域「{text}」不存在")
        return row.id
    row = (
        db.query(SalesTerritory)
        .filter(SalesTerritory.tenant_id == tenant_id, SalesTerritory.name == text)
        .first()
    )
    if not row:
        raise HTTPException(status_code=400, detail=f"销售区域「{text}」不存在")
    return row.id


def _parse_boolish(val) -> bool:
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "是", "启用", "开", "含税"):
        return True
    if s in ("0", "false", "no", "n", "否", "停用", "关", "不含税"):
        return False
    raise HTTPException(status_code=400, detail=f"布尔值无效: {val}（请填 是/否）")


def _parse_tax_rate(val) -> float:
    """解析税率：支持 13、13%、Excel 百分比单元格 0.13。"""
    if val is None or val == "":
        raise HTTPException(status_code=400, detail="税率不能为空")
    if isinstance(val, bool):
        raise HTTPException(status_code=400, detail="税率须为数字")
    if isinstance(val, (int, float)):
        rate = float(val)
    else:
        s = str(val).strip().replace("%", "").replace("％", "")
        if not s:
            raise HTTPException(status_code=400, detail="税率不能为空")
        try:
            rate = float(s)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"税率须为数字: {val}") from e
    # Excel 百分比格式常读成 0.13
    if 0 < rate < 1:
        rate = round(rate * 100, 4)
    if rate < 0 or rate > 100:
        raise HTTPException(status_code=400, detail=f"税率须在 0～100 之间: {val}")
    return round(rate, 2)


def _cell_to_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return "是" if val else "否"
    if isinstance(val, float):
        rounded = round(val)
        if abs(val - rounded) < 1e-9:
            return str(int(rounded))
        return str(val).strip()
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _decode_csv_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="无法识别 CSV 编码，请另存为 UTF-8 或 GBK")


def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    text = _decode_csv_text(content)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV 缺少表头")
    columns = [c.strip() for c in reader.fieldnames if c and c.strip()]
    rows: list[dict] = []
    for i, row in enumerate(reader, start=2):
        if i - 1 > MAX_IMPORT_ROWS:
            raise HTTPException(status_code=400, detail=f"超过最大行数 {MAX_IMPORT_ROWS}")
        cleaned = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        if any(v for v in cleaned.values()):
            rows.append(cleaned)
    return columns, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    try:
        import openpyxl
    except ImportError as e:
        raise HTTPException(status_code=400, detail="XLSX 解析需要 openpyxl") from e
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    # 优先读名为「数据」的工作表，否则取第一个非「说明」表
    ws = None
    for name in wb.sheetnames:
        if name.strip() in ("数据", "导入数据", "Sheet1"):
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if name.strip() not in ("说明", "填写说明", "readme"):
                ws = wb[name]
                break
    if ws is None:
        ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise HTTPException(status_code=400, detail="Excel 缺少表头")
    columns = [str(c).strip() for c in header if c is not None and str(c).strip()]
    if not columns:
        raise HTTPException(status_code=400, detail="Excel 缺少表头")
    rows: list[dict] = []
    for i, cells in enumerate(rows_iter, start=2):
        if i - 1 > MAX_IMPORT_ROWS:
            raise HTTPException(status_code=400, detail=f"超过最大行数 {MAX_IMPORT_ROWS}")
        data = {}
        for idx, col in enumerate(columns):
            val = cells[idx] if cells and idx < len(cells) else None
            if val is not None and val != "":
                data[col] = _cell_to_str(val)
        if any(data.values()):
            rows.append(data)
    wb.close()
    return columns, rows


def parse_upload_file(filename: str, content: bytes) -> tuple[list[str], list[dict]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return _parse_csv(content)
    if lower.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise HTTPException(status_code=400, detail="仅支持 Excel（.xlsx）或 CSV（UTF-8）")


def _import_header_label(label: str, *, is_required: bool) -> str:
    return f"{label}*" if is_required else label


def _strip_import_header(col: str) -> str:
    return col.rstrip("*").strip()


def suggest_mapping(
    db: Session, tenant_id: UUID, entity_type: str, columns: list[str]
) -> dict[str, str]:
    if entity_type == "product_spec_model":
        label_map: dict[str, str] = {}
        key_set = set(SPEC_MODEL_DB_KEYS)
        for key, label, required in SPEC_MODEL_IMPORT_FIELDS:
            label_map[label] = key
            if required:
                label_map[_import_header_label(label, is_required=True)] = key
        out: dict[str, str] = {}
        for col in columns:
            bare = _strip_import_header(col)
            if col in key_set:
                out[col] = col
            elif bare in key_set:
                out[col] = bare
            elif col in label_map:
                out[col] = label_map[col]
            elif bare in label_map:
                out[col] = label_map[bare]
            elif bare in SPEC_MODEL_COLUMN_ALIASES:
                out[col] = SPEC_MODEL_COLUMN_ALIASES[bare]
        return out

    ensure_entity_schema(db, tenant_id, entity_type)
    fields = list_active_fields(db, tenant_id, entity_type)
    label_map: dict[str, str] = {}
    for f in fields:
        label_map[f.label] = f.field_key
        if _import_field_required(entity_type, f.field_key, schema_required=bool(f.is_required)):
            label_map[_import_header_label(f.label, is_required=True)] = f.field_key
    key_set = {f.field_key for f in fields}
    out: dict[str, str] = {}
    for col in columns:
        bare = _strip_import_header(col)
        if col in key_set:
            out[col] = col
        elif bare in key_set:
            out[col] = bare
        elif col in label_map:
            out[col] = label_map[col]
        elif bare in label_map:
            out[col] = label_map[bare]
        elif bare in IMPORT_COLUMN_ALIASES and IMPORT_COLUMN_ALIASES[bare] in key_set:
            out[col] = IMPORT_COLUMN_ALIASES[bare]
    return out


def _template_headers(db: Session, tenant_id: UUID, entity_type: str) -> list[str]:
    if entity_type == "product_spec_model":
        return [
            _import_header_label(label, is_required=required)
            for _key, label, required in SPEC_MODEL_IMPORT_FIELDS
        ]

    ensure_entity_schema(db, tenant_id, entity_type)
    fields = list_active_fields(db, tenant_id, entity_type)
    skip = {"created_by_user_id", "created_at", "updated_at", "converted_customer_id", "converted_from_lead_id"}
    headers: list[str] = []
    for f in fields:
        if f.field_key in skip:
            continue
        # 产品编码由系统自动生成，导入模板不包含该列
        if entity_type == "product" and f.field_key == "code":
            continue
        if f.field_type in ("user_ref", "territory_ref", "ref"):
            # 产品分类/规格型号：模板填名称，导入时解析为 ID
            # 线索销售区域：模板填地区名称
            allow_name_ref = (entity_type == "product" and f.field_key in PRODUCT_REF_NAME_KEYS) or (
                entity_type == "lead" and f.field_key == "territory_id"
            )
            if not allow_name_ref:
                continue
        if not f.is_active:
            continue
        headers.append(
            _import_header_label(
                f.label,
                is_required=_import_field_required(entity_type, f.field_key, schema_required=bool(f.is_required)),
            )
        )
    return headers


def _product_sample_row(headers: list[str]) -> list[str]:
    """产品模板示例行：含税标价 113 + 税率 13%，便于对照价税语义。"""
    sample_by_label = {
        "产品名称": "价税示例产品（可删）",
        "产品名称*": "价税示例产品（可删）",
        "单位": "",
        "标价": "113",
        "成本价": "80",
        "默认税率%": "13",
        "标价含税": "是",
        "分类": "",
        "规格型号": "",
        "启用": "是",
        "描述": "标价含税时选入报价会拆未税单价；导入前可删本行",
    }
    return [sample_by_label.get(h, "") for h in headers]


def _spec_model_sample_row(headers: list[str]) -> list[str]:
    sample_by_label = {
        "名称": "标准版（可删）",
        "名称*": "标准版（可删）",
        "编码": "STD-001",
        "说明": "示例规格型号",
        "排序": "0",
        "启用": "是",
    }
    return [sample_by_label.get(h, "") for h in headers]


def build_template_csv(db: Session, tenant_id: UUID, entity_type: str) -> str:
    headers = _template_headers(db, tenant_id, entity_type)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    if entity_type == "product":
        writer.writerow(_product_sample_row(headers))
    elif entity_type == "product_spec_model":
        writer.writerow(_spec_model_sample_row(headers))
    return buf.getvalue()


def build_template_xlsx(db: Session, tenant_id: UUID, entity_type: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise HTTPException(status_code=500, detail="生成 Excel 模板需要 openpyxl") from e

    headers = _template_headers(db, tenant_id, entity_type)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据"
    header_fill = PatternFill("solid", fgColor="E8F0FE")
    header_font = Font(bold=True)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(12, min(28, len(title) * 2 + 2))
    if entity_type == "product":
        for col_idx, val in enumerate(_product_sample_row(headers), start=1):
            ws.cell(row=2, column=col_idx, value=val if val != "" else None)
    elif entity_type == "product_spec_model":
        for col_idx, val in enumerate(_spec_model_sample_row(headers), start=1):
            ws.cell(row=2, column=col_idx, value=val if val != "" else None)

    tip = wb.create_sheet("填写说明", 1)
    tips = [
        "导入说明",
        "1. 请在「数据」工作表填写；勿改表头名称（带 * 为必填）。",
        "2. 支持上传本模板（.xlsx），也兼容 CSV（UTF-8）。",
        "3. 最大行数 5000；导入后按列映射确认即可。",
    ]
    if entity_type == "product":
        tips.extend(
            [
                "4. 产品编码由系统自动生成，模板不含该列；无需填写。",
                "5. 【默认税率%】填 0～100 的数字，如 13、9、6；也可写 13%。",
                "6. Excel 若单元格设为「百分比」格式（显示 13%），系统会按 13 识别。",
                "7. 【标价含税】填 是/否（或 true/false）。为「是」时，选入报价会按税率拆未税单价。",
                "8. 分类、规格型号、单位请填系统中已存在的名称（非 ID）。",
                "9. 第 2 行为示例，导入前可删除。",
            ]
        )
    elif entity_type == "product_spec_model":
        tips.extend(
            [
                "4. 【名称】租户内不可重复；重复行可选「跳过」或「更新已有」。",
                "5. 【编码】可选；若填写则租户内不可重复。",
                "6. 【启用】填 是/否。",
                "7. 第 2 行为示例，导入前可删除。",
            ]
        )
    else:
        tips.append("4. 重复数据按导入选项处理（跳过/更新）。")
        if entity_type == "lead":
            tips.append(
                "5. 【销售区域】填系统中已有地区名称（如华东）；留空则使用导入人在销售组织中的主地区。"
            )
    for i, line in enumerate(tips, start=1):
        tip.cell(row=i, column=1, value=line)
        if i == 1:
            tip.cell(row=i, column=1).font = Font(bold=True, size=14)
    tip.column_dimensions["A"].width = 72

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_import_template(
    db: Session, tenant_id: UUID, entity_type: str, *, fmt: str = "xlsx"
) -> tuple[bytes, str, str]:
    """返回 (content, media_type, filename)。"""
    fmt = (fmt or "xlsx").lower().strip()
    if fmt not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="模板格式仅支持 xlsx 或 csv")
    if fmt == "csv":
        text = build_template_csv(db, tenant_id, entity_type)
        return (
            text.encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            f"{entity_type}_import_template.csv",
        )
    content = build_template_xlsx(db, tenant_id, entity_type)
    return (
        content,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"{entity_type}_import_template.xlsx",
    )


def create_import_job(
    db: Session,
    ctx: TenantContext,
    entity_type: str,
    filename: str,
    content: bytes,
) -> tuple[CrmImportJob, list[str], dict[str, str]]:
    if entity_type not in IMPORT_ENTITY_TYPES:
        raise HTTPException(status_code=400, detail=f"不支持导入实体: {entity_type}")

    columns, rows = parse_upload_file(filename, content)
    suggested = suggest_mapping(db, ctx.tenant_id, entity_type, columns)
    job = CrmImportJob(
        tenant_id=ctx.tenant_id,
        entity_type=entity_type,
        status="draft",
        file_name=filename,
        file_storage_path="",
        mapping=suggested or {},
        options=_default_options(entity_type),
        created_by_user_id=ctx.user.id,
        total_rows=len(rows),
        columns=columns,
    )
    db.add(job)
    db.flush()

    dest = _import_dir(ctx.tenant_id, job.id) / filename
    dest.write_bytes(content)
    job.file_storage_path = str(dest)
    db.commit()
    db.refresh(job)

    job._parsed_rows = rows  # type: ignore[attr-defined]
    return job, columns, suggested


def get_job(db: Session, tenant_id: UUID, job_id: UUID) -> CrmImportJob | None:
    return (
        db.query(CrmImportJob)
        .filter(CrmImportJob.id == job_id, CrmImportJob.tenant_id == tenant_id)
        .first()
    )


def _load_rows(job: CrmImportJob) -> list[dict]:
    path = Path(job.file_storage_path)
    if not path.is_file():
        raise HTTPException(status_code=400, detail="导入文件不存在")
    _, rows = parse_upload_file(path.name, path.read_bytes())
    return rows


def _perm_set(ctx: TenantContext, db: Session) -> set[str]:
    membership = get_membership(db, ctx.user.id, ctx.tenant_id)
    if not membership or not membership.role:
        return set()
    return {p.permission_code for p in membership.role.permissions}


def _row_to_payload(
    db: Session,
    ctx: TenantContext,
    entity_type: str,
    raw: dict,
    mapping: dict,
    options: dict,
) -> tuple[dict, dict, str | None]:
    """返回 db_fields, extra_data, error。"""
    db_keys = _db_keys_for(entity_type)
    db_fields: dict = {}
    extra: dict = {}
    for col, field_key in mapping.items():
        if col not in raw:
            continue
        val = raw.get(col)
        if val is None or val == "":
            continue
        if field_key in db_keys:
            db_fields[field_key] = val

    if entity_type == "product_spec_model":
        if not db_fields.get("name") or not str(db_fields.get("name")).strip():
            return {}, {}, "规格型号名称不能为空"
        db_fields["name"] = str(db_fields["name"]).strip()
        if db_fields.get("code") and str(db_fields.get("code")).strip():
            db_fields["code"] = str(db_fields["code"]).strip()
        else:
            db_fields.pop("code", None)
        try:
            if "sort_order" in db_fields:
                db_fields["sort_order"] = int(float(db_fields["sort_order"]))
            if "is_active" in db_fields:
                db_fields["is_active"] = _parse_boolish(db_fields["is_active"])
            if "description" in db_fields:
                db_fields["description"] = str(db_fields["description"]).strip()
        except HTTPException as e:
            return {}, {}, str(e.detail)
        except (TypeError, ValueError):
            return {}, {}, "排序须为数字"
        return db_fields, {}, None

    ensure_entity_schema(db, ctx.tenant_id, entity_type)
    field_map = {f.field_key: f for f in list_active_fields(db, ctx.tenant_id, entity_type)}

    db_fields = {}
    extra = {}
    for col, field_key in mapping.items():
        if col not in raw:
            continue
        val = raw.get(col)
        if val is None or val == "":
            continue
        fdef = field_map.get(field_key)
        if not fdef or not fdef.is_active:
            continue
        if field_key in db_keys:
            db_fields[field_key] = val
        else:
            extra[field_key] = val

    if entity_type == "product":
        if not db_fields.get("name") or not str(db_fields.get("name")).strip():
            return {}, {}, "产品名称不能为空"
        db_fields["name"] = str(db_fields["name"]).strip()
        if db_fields.get("code") and str(db_fields.get("code")).strip():
            db_fields["code"] = str(db_fields["code"]).strip()
        else:
            # 空编码 → 交由 create_product 自动流水号
            db_fields.pop("code", None)
        try:
            if "list_price" in db_fields:
                db_fields["list_price"] = float(db_fields["list_price"])
            if "cost_price" in db_fields:
                db_fields["cost_price"] = float(db_fields["cost_price"])
            if "default_tax_rate" in db_fields and db_fields["default_tax_rate"] not in (None, ""):
                db_fields["default_tax_rate"] = _parse_tax_rate(db_fields["default_tax_rate"])
            if "price_includes_tax" in db_fields:
                db_fields["price_includes_tax"] = _parse_boolish(db_fields["price_includes_tax"])
            if "is_active" in db_fields:
                db_fields["is_active"] = _parse_boolish(db_fields["is_active"])
            if "unit" in db_fields:
                db_fields["unit"] = _resolve_product_unit_name(db, ctx.tenant_id, str(db_fields["unit"]))
            if "category_id" in db_fields:
                db_fields["category_id"] = _resolve_product_category_id(
                    db, ctx.tenant_id, str(db_fields["category_id"])
                )
            if "spec_model_id" in db_fields:
                db_fields["spec_model_id"] = resolve_spec_model_id_by_name(
                    db, ctx.tenant_id, str(db_fields["spec_model_id"])
                )
        except HTTPException as e:
            return {}, {}, str(e.detail)
        except (TypeError, ValueError):
            return {}, {}, "价格/税率字段须为数字"

        try:
            extra = validate_extra_data(db, ctx.tenant_id, entity_type, extra, is_create=True)
        except HTTPException as e:
            return {}, {}, str(e.detail)
        return db_fields, extra, None

    if not db_fields.get("company_name"):
        return {}, {}, "公司名称不能为空"

    if entity_type == "lead":
        if not db_fields.get("contact_name") or not str(db_fields.get("contact_name")).strip():
            return {}, {}, "联系人姓名不能为空"
        db_fields.setdefault("status", options.get("default_status") or "待跟进")
        db_fields.setdefault("source", options.get("default_source") or "导入")
        try:
            validate_lead_status(db_fields["status"])
        except ValueError as e:
            return {}, {}, str(e)
        mobile, mobile_err = validate_lead_mobile_value(db_fields.get("mobile"), required=True)
        if mobile_err:
            return {}, {}, mobile_err
        db_fields["mobile"] = mobile
        try:
            raw_tid = db_fields.get("territory_id")
            if raw_tid in (None, ""):
                raw_tid = options.get("default_territory_id")
            resolved: UUID | None = None
            if raw_tid not in (None, ""):
                resolved = _resolve_territory_id(db, ctx.tenant_id, str(raw_tid))
            territory_id, manager_user_id = apply_creator_org_defaults(
                db, ctx, territory_id=resolved
            )
            if territory_id is None:
                return (
                    {},
                    {},
                    "销售区域不能为空：请填写地区名称，或在「设置 → 销售组织」中配置本人主地区",
                )
            db_fields["territory_id"] = territory_id
            db_fields["manager_user_id"] = manager_user_id
        except HTTPException as e:
            return {}, {}, str(e.detail)
    else:
        db_fields.setdefault("status", options.get("default_status") or "潜在")
        try:
            validate_customer_status(db_fields["status"])
        except ValueError as e:
            return {}, {}, str(e)

        mobile = db_fields.get("mobile")
        if mobile:
            mobile_str = str(mobile).strip()
            if _PHONE_RE.match(mobile_str):
                db_fields["mobile"] = mobile_str
            else:
                db_fields.pop("mobile", None)

    owner = options.get("default_owner_user_id") or str(ctx.user.id)
    db_fields["owner_user_id"] = UUID(str(owner))

    try:
        extra = validate_extra_data(db, ctx.tenant_id, entity_type, extra, is_create=True)
    except HTTPException as e:
        return {}, {}, str(e.detail)

    return db_fields, extra, None


def _is_admin(ctx: TenantContext) -> bool:
    return ctx.membership.role.is_system and ctx.membership.role.code == SYSTEM_ROLE_ADMIN


def list_import_jobs(
    db: Session,
    ctx: TenantContext,
    *,
    page: int = 1,
    page_size: int = 20,
    entity_type: str | None = None,
) -> tuple[list[CrmImportJob], int]:
    query = db.query(CrmImportJob).filter(CrmImportJob.tenant_id == ctx.tenant_id)
    if not _is_admin(ctx):
        query = query.filter(CrmImportJob.created_by_user_id == ctx.user.id)
    if entity_type:
        if entity_type not in IMPORT_ENTITY_TYPES:
            raise HTTPException(status_code=400, detail=f"不支持的 entity_type: {entity_type}")
        query = query.filter(CrmImportJob.entity_type == entity_type)
    total = query.count()
    items = (
        query.order_by(CrmImportJob.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return items, total


def _apply_update_to_entity(
    db: Session,
    ctx: TenantContext,
    existing: Lead | Customer | Product | ProductSpecModel,
    entity_type: str,
    db_fields: dict,
    extra: dict,
) -> None:
    if entity_type == "product_spec_model":
        assert isinstance(existing, ProductSpecModel)
        payload = {k: v for k, v in db_fields.items() if k in SPEC_MODEL_DB_KEYS}
        update_spec_model(db, ctx, existing, ProductSpecModelUpdate(**payload))
        return

    if entity_type == "product":
        assert isinstance(existing, Product)
        payload = {k: v for k, v in db_fields.items() if k in PRODUCT_DB_KEYS}
        if extra:
            payload["extra_data"] = extra
        update_product(db, ctx, existing, ProductUpdate(**payload))
        return

    if entity_type == "lead":
        assert_can_view_lead(
            ctx,
            db,
            existing.owner_user_id,
            existing.territory_id,
            created_by_user_id=existing.created_by_user_id,
            manager_user_id=getattr(existing, "manager_user_id", None),
            pool_id=getattr(existing, "pool_id", None),
        )
        if "company_name" in db_fields:
            existing.company_name = db_fields["company_name"]
        for key in ("contact_name", "mobile", "phone", "email", "source", "remark"):
            if key in db_fields:
                setattr(existing, key, db_fields[key])
        if "status" in db_fields:
            validate_lead_status(db_fields["status"])
            existing.status = db_fields["status"]
        if "territory_id" in db_fields and db_fields["territory_id"]:
            existing.territory_id = db_fields["territory_id"]
        if "manager_user_id" in db_fields:
            existing.manager_user_id = db_fields["manager_user_id"]
    else:
        assert_can_view_customer(
            ctx,
            db,
            existing.owner_user_id,
            existing.territory_id,
            created_by_user_id=existing.created_by_user_id,
            manager_user_id=getattr(existing, "manager_user_id", None),
            pool_id=getattr(existing, "pool_id", None),
        )
        if "company_name" in db_fields:
            existing.company_name = db_fields["company_name"]
        for key in ("mobile", "phone", "email", "remark"):
            if key in db_fields:
                setattr(existing, key, db_fields[key])
        if "status" in db_fields:
            validate_customer_status(db_fields["status"])
            existing.status = db_fields["status"]
    if extra:
        merged = dict(existing.extra_data or {})
        merged.update(extra)
        existing.extra_data = validate_extra_data(db, ctx.tenant_id, entity_type, merged)


def _find_duplicate(
    db: Session, tenant_id: UUID, entity_type: str, key: str, value
) -> Lead | Customer | Product | ProductSpecModel | None:
    if not value:
        return None
    if entity_type == "lead":
        if key == "mobile":
            return (
                db.query(Lead)
                .filter(Lead.tenant_id == tenant_id, Lead.mobile == value, Lead.deleted_at.is_(None))
                .first()
            )
    elif entity_type == "customer":
        if key == "mobile":
            return (
                db.query(Customer)
                .filter(Customer.tenant_id == tenant_id, Customer.mobile == value, Customer.deleted_at.is_(None))
                .first()
            )
    elif entity_type == "product":
        if key == "code":
            return (
                db.query(Product)
                .filter(
                    Product.tenant_id == tenant_id,
                    Product.code == str(value).strip(),
                    Product.deleted_at.is_(None),
                )
                .first()
            )
    elif entity_type == "product_spec_model":
        if key == "name":
            return (
                db.query(ProductSpecModel)
                .filter(
                    ProductSpecModel.tenant_id == tenant_id,
                    ProductSpecModel.name == str(value).strip(),
                )
                .first()
            )
        if key == "code":
            return (
                db.query(ProductSpecModel)
                .filter(
                    ProductSpecModel.tenant_id == tenant_id,
                    ProductSpecModel.code == str(value).strip(),
                )
                .first()
            )
    return None


def preview_job(db: Session, ctx: TenantContext, job: CrmImportJob) -> dict:
    rows = _load_rows(job)
    mapping = job.mapping or {}
    options = {**_default_options(job.entity_type), **(job.options or {})}
    preview_rows: list[dict] = []
    error_count = 0
    ok_count = 0

    for i, raw in enumerate(rows[:20], start=1):
        db_fields, extra, err = _row_to_payload(db, ctx, job.entity_type, raw, mapping, options)
        if err:
            preview_rows.append(
                {"row_number": i, "status": "preview_error", "error_message": err, "data": raw}
            )
            error_count += 1
        else:
            dup_key = options.get("duplicate_key")
            dup = None
            if dup_key and db_fields.get(dup_key):
                dup = _find_duplicate(db, ctx.tenant_id, job.entity_type, dup_key, db_fields.get(dup_key))
            if dup and options.get("on_duplicate") == "skip":
                preview_rows.append(
                    {
                        "row_number": i,
                        "status": "preview_ok",
                        "error_message": "将跳过重复",
                        "data": raw,
                    }
                )
            elif dup and options.get("on_duplicate") == "update":
                preview_rows.append(
                    {
                        "row_number": i,
                        "status": "preview_ok",
                        "error_message": "将更新已有记录",
                        "data": raw,
                    }
                )
            else:
                preview_rows.append({"row_number": i, "status": "preview_ok", "data": raw})
            ok_count += 1

    job.status = "previewing"
    db.commit()
    return {"preview_rows": preview_rows, "error_count": error_count, "ok_count": ok_count}


def run_import(db: Session, ctx: TenantContext, job: CrmImportJob) -> CrmImportJob:
    if job.status == "completed":
        return job
    rows = _load_rows(job)
    mapping = job.mapping or {}
    options = {**_default_options(job.entity_type), **(job.options or {})}

    db.query(CrmImportRow).filter(CrmImportRow.job_id == job.id).delete()
    job.status = "importing"
    job.started_at = datetime.now(timezone.utc)
    job.success_count = 0
    job.skip_count = 0
    job.error_count = 0
    job.total_rows = len(rows)
    db.commit()

    perms = _perm_set(ctx, db)
    if job.entity_type == "product":
        assign_perm = ""
        create_perm = "crm.product.manage"
    elif job.entity_type == "product_spec_model":
        assign_perm = ""
        create_perm = "crm.product.manage"
    elif job.entity_type == "lead":
        assign_perm = "crm.lead.assign"
        create_perm = "crm.lead.create"
    else:
        assign_perm = "crm.customer.assign"
        create_perm = "crm.customer.create"
    if create_perm not in perms:
        job.status = "failed"
        db.commit()
        raise HTTPException(status_code=403, detail="无创建权限")

    for i, raw in enumerate(rows, start=1):
        db_fields, extra, err = _row_to_payload(db, ctx, job.entity_type, raw, mapping, options)
        if err:
            job.error_count += 1
            db.add(
                CrmImportRow(
                    job_id=job.id,
                    row_number=i,
                    raw_data=raw,
                    status="error",
                    error_message=err,
                )
            )
            continue

        owner_id = db_fields.pop("owner_user_id", ctx.user.id) if job.entity_type not in (
            "product",
            "product_spec_model",
        ) else None
        if owner_id is not None and owner_id != ctx.user.id and assign_perm and assign_perm not in perms:
            owner_id = ctx.user.id

        dup_key = options.get("duplicate_key")
        dup_val = db_fields.get(dup_key) if dup_key else None
        existing = _find_duplicate(db, ctx.tenant_id, job.entity_type, dup_key or "", dup_val) if dup_val else None

        if existing and options.get("on_duplicate") == "skip":
            job.skip_count += 1
            db.add(
                CrmImportRow(
                    job_id=job.id,
                    row_number=i,
                    raw_data=raw,
                    status="skip",
                    error_message="重复跳过",
                    target_id=existing.id,
                )
            )
            continue

        if existing and options.get("on_duplicate") == "update":
            try:
                update_fields = dict(db_fields)
                update_fields.pop("owner_user_id", None)
                _apply_update_to_entity(db, ctx, existing, job.entity_type, update_fields, extra)
                job.success_count += 1
                db.add(
                    CrmImportRow(
                        job_id=job.id,
                        row_number=i,
                        raw_data=raw,
                        status="success",
                        error_message="已更新",
                        target_id=existing.id,
                    )
                )
            except HTTPException as e:
                job.error_count += 1
                db.add(
                    CrmImportRow(
                        job_id=job.id,
                        row_number=i,
                        raw_data=raw,
                        status="error",
                        error_message=str(e.detail),
                    )
                )
            except Exception as e:
                job.error_count += 1
                db.add(
                    CrmImportRow(
                        job_id=job.id,
                        row_number=i,
                        raw_data=raw,
                        status="error",
                        error_message=str(e),
                    )
                )
            continue

        try:
            if job.entity_type == "product":
                create_payload = {
                    "name": db_fields["name"],
                    "code": db_fields.get("code"),
                    "unit": db_fields.get("unit"),
                    "list_price": db_fields.get("list_price", 0) or 0,
                    "cost_price": db_fields.get("cost_price"),
                    "default_tax_rate": db_fields.get("default_tax_rate"),
                    "price_includes_tax": db_fields.get("price_includes_tax", False),
                    "category_id": db_fields.get("category_id"),
                    "spec_model_id": db_fields.get("spec_model_id"),
                    "is_active": db_fields.get("is_active", True),
                    "description": db_fields.get("description"),
                    "extra_data": extra,
                }
                product = create_product(db, ctx, ProductCreate(**create_payload))
                target_id = product.id
            elif job.entity_type == "product_spec_model":
                spec = create_spec_model(
                    db,
                    ctx,
                    ProductSpecModelCreate(
                        name=db_fields["name"],
                        code=db_fields.get("code"),
                        description=db_fields.get("description"),
                        sort_order=db_fields.get("sort_order", 0),
                        is_active=db_fields.get("is_active", True),
                    ),
                )
                target_id = spec.id
            elif job.entity_type == "lead":
                lead = Lead(
                    tenant_id=ctx.tenant_id,
                    lead_number=generate_number(db, ctx.tenant_id, "lead"),
                    company_name=db_fields.pop("company_name"),
                    contact_name=db_fields.get("contact_name"),
                    mobile=db_fields.get("mobile"),
                    phone=db_fields.get("phone"),
                    email=db_fields.get("email"),
                    source=db_fields.get("source"),
                    status=db_fields.get("status", "待跟进"),
                    remark=db_fields.get("remark"),
                    owner_user_id=owner_id,
                    territory_id=db_fields.get("territory_id"),
                    manager_user_id=db_fields.get("manager_user_id"),
                    extra_data=extra,
                    created_by_user_id=ctx.user.id,
                )
                db.add(lead)
                db.flush()
                apply_assignment_rules(db, ctx, lead)
                target_id = lead.id
            else:
                customer = Customer(
                    tenant_id=ctx.tenant_id,
                    company_name=db_fields.pop("company_name"),
                    mobile=db_fields.get("mobile"),
                    phone=db_fields.get("phone"),
                    email=db_fields.get("email"),
                    status=db_fields.get("status", "潜在"),
                    remark=db_fields.get("remark"),
                    owner_user_id=owner_id,
                    extra_data=extra,
                    created_by_user_id=ctx.user.id,
                )
                db.add(customer)
                db.flush()
                target_id = customer.id

            job.success_count += 1
            db.add(
                CrmImportRow(
                    job_id=job.id,
                    row_number=i,
                    raw_data=raw,
                    status="success",
                    target_id=target_id,
                )
            )
        except Exception as e:
            job.error_count += 1
            detail = str(e.detail) if isinstance(e, HTTPException) else str(e)
            db.add(
                CrmImportRow(
                    job_id=job.id,
                    row_number=i,
                    raw_data=raw,
                    status="error",
                    error_message=detail,
                )
            )

    job.status = "completed"
    job.completed_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(job)
    return job


def build_errors_csv(db: Session, job: CrmImportJob) -> str:
    rows = (
        db.query(CrmImportRow)
        .filter(CrmImportRow.job_id == job.id, CrmImportRow.status.in_(("error", "skip")))
        .order_by(CrmImportRow.row_number)
        .all()
    )
    if not rows:
        return "row_number,error_message\n"
    all_keys: list[str] = []
    for r in rows:
        for k in r.raw_data:
            if k not in all_keys:
                all_keys.append(k)
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=["row_number", "error_message", "status", *all_keys])
    writer.writeheader()
    for r in rows:
        line = {"row_number": r.row_number, "error_message": r.error_message or "", "status": r.status}
        line.update(r.raw_data)
        writer.writerow(line)
    return buf.getvalue()


def update_job_mapping(
    db: Session, job: CrmImportJob, mapping: dict, options: dict | None
) -> CrmImportJob:
    job.mapping = mapping
    if options:
        merged = {**_default_options(job.entity_type), **job.options, **options}
        job.options = merged
    job.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(job)
    return job
