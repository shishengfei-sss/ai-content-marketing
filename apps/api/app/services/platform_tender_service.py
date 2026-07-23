"""平台公共招标线索池 L1 服务。"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from uuid import UUID

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.models import User
from app.models.tender import PlatformTenderLead
from app.schemas.platform_tender import (
    PlatformTenderExcelConfirmOut,
    PlatformTenderExcelPreviewOut,
    PlatformTenderExcelPreviewRow,
    PlatformTenderLeadCreate,
    PlatformTenderLeadUpdate,
)

EXCEL_HEADERS = [
    "采购方",
    "行业",
    "地区",
    "产品/标的",
    "数量",
    "预算下限",
    "预算上限",
    "截止日期",
    "联系人",
    "联系电话",
    "原文链接",
    "摘要",
    "项目编号",
    "发布时间",
    "采购方式",
    "代理单位",
    "采购人地址",
    "品目分类",
    "开标日期",
    "面向中小企业",
    "资格要求摘要",
    "最高限价",
]


def _assert_source_url(*, source_url: str | None, has_source_document: bool, status_value: str) -> None:
    if has_source_document and not source_url:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="有原文时必须填写 source_url（原文链接）",
        )
    if status_value == "published" and not source_url:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="发布到公共池须填写原文链接 source_url",
        )


def list_leads(
    db: Session,
    *,
    status_filter: str | None,
    q: str | None,
    page: int,
    page_size: int,
    region: str | None = None,
    industry: str | None = None,
    category: str | None = None,
    procurement_method: str | None = None,
    agent_name: str | None = None,
    project_no: str | None = None,
    sme_preference: bool | None = None,
    deadline_from: date | None = None,
    deadline_to: date | None = None,
    published_from: date | None = None,
    published_to: date | None = None,
) -> tuple[list[PlatformTenderLead], int]:
    query = db.query(PlatformTenderLead)
    if status_filter:
        query = query.filter(PlatformTenderLead.status == status_filter)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(
                PlatformTenderLead.buyer_name.ilike(like),
                PlatformTenderLead.product_name.ilike(like),
                PlatformTenderLead.region.ilike(like),
                PlatformTenderLead.industry.ilike(like),
                PlatformTenderLead.project_no.ilike(like),
                PlatformTenderLead.agent_name.ilike(like),
                PlatformTenderLead.category.ilike(like),
                PlatformTenderLead.summary.ilike(like),
                PlatformTenderLead.buyer_address.ilike(like),
            )
        )
    if region and region.strip():
        query = query.filter(PlatformTenderLead.region.ilike(f"%{region.strip()}%"))
    if industry and industry.strip():
        query = query.filter(PlatformTenderLead.industry.ilike(f"%{industry.strip()}%"))
    if category and category.strip():
        query = query.filter(PlatformTenderLead.category.ilike(f"%{category.strip()}%"))
    if procurement_method and procurement_method.strip():
        query = query.filter(
            PlatformTenderLead.procurement_method.ilike(f"%{procurement_method.strip()}%")
        )
    if agent_name and agent_name.strip():
        query = query.filter(PlatformTenderLead.agent_name.ilike(f"%{agent_name.strip()}%"))
    if project_no and project_no.strip():
        query = query.filter(PlatformTenderLead.project_no.ilike(f"%{project_no.strip()}%"))
    if sme_preference is not None:
        query = query.filter(PlatformTenderLead.sme_preference.is_(sme_preference))
    if deadline_from:
        query = query.filter(PlatformTenderLead.deadline >= deadline_from)
    if deadline_to:
        query = query.filter(PlatformTenderLead.deadline <= deadline_to)
    if published_from:
        query = query.filter(PlatformTenderLead.published_at >= published_from)
    if published_to:
        query = query.filter(PlatformTenderLead.published_at <= published_to)
    total = query.count()
    items = (
        query.order_by(PlatformTenderLead.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return items, total


def get_lead(db: Session, lead_id: UUID) -> PlatformTenderLead:
    row = db.query(PlatformTenderLead).filter(PlatformTenderLead.id == lead_id).first()
    if not row:
        raise HTTPException(status_code=404, detail="招标线索不存在")
    return row


def create_lead(db: Session, admin: User, data: PlatformTenderLeadCreate) -> PlatformTenderLead:
    _assert_source_url(
        source_url=data.source_url,
        has_source_document=data.has_source_document,
        status_value=data.status,
    )
    row = PlatformTenderLead(
        buyer_name=data.buyer_name.strip(),
        industry=data.industry,
        region=data.region,
        product_name=data.product_name,
        quantity=data.quantity,
        budget_min=data.budget_min,
        budget_max=data.budget_max,
        deadline=data.deadline,
        contact_name=data.contact_name,
        contact_phone=data.contact_phone,
        source_url=data.source_url,
        summary=data.summary,
        project_no=data.project_no,
        published_at=data.published_at,
        procurement_method=data.procurement_method,
        agent_name=data.agent_name,
        buyer_address=data.buyer_address,
        category=data.category,
        bid_open_date=data.bid_open_date,
        sme_preference=data.sme_preference,
        qualification_summary=data.qualification_summary,
        max_price_limit=data.max_price_limit,
        source_channel=data.source_channel,
        status=data.status,
        created_by=admin.id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return row


def update_lead(db: Session, row: PlatformTenderLead, data: PlatformTenderLeadUpdate) -> PlatformTenderLead:
    payload = data.model_dump(exclude_unset=True)
    has_source = payload.pop("has_source_document", None)
    for k, v in payload.items():
        if k == "buyer_name" and isinstance(v, str):
            setattr(row, k, v.strip())
        else:
            setattr(row, k, v)

    if has_source is True and not row.source_url:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="有原文时必须填写 source_url（原文链接）",
        )
    if row.status == "published" and not row.source_url:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="发布到公共池须填写原文链接 source_url",
        )

    db.commit()
    db.refresh(row)
    return row


def set_status(db: Session, row: PlatformTenderLead, new_status: str) -> PlatformTenderLead:
    if new_status not in ("draft", "published", "unpublished"):
        raise HTTPException(status_code=400, detail="非法状态")
    if new_status == "published":
        _assert_source_url(source_url=row.source_url, has_source_document=True, status_value="published")
    row.status = new_status
    db.commit()
    db.refresh(row)
    if new_status == "published":
        try:
            from app.services.tender_match_service import rematch_platform_lead

            rematch_platform_lead(db, row.id)
        except Exception:
            # 匹配失败不阻断发布
            pass
    return row


def delete_lead(db: Session, row: PlatformTenderLead) -> None:
    db.delete(row)
    db.commit()


def build_excel_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "招标线索"
    ws.append(EXCEL_HEADERS)
    ws.append(
        [
            "示例市第一人民医院",
            "医疗",
            "浙江杭州",
            "离心泵",
            "2台",
            50000,
            120000,
            "2026-12-31",
            "张工",
            "13800000000",
            "https://example.com/tender/demo",
            "示例行，导入前请删除",
            "SZCG2026000001",
            "2026-07-01",
            "公开招标",
            "示例招标代理有限公司",
            "杭州市西湖区示例路1号",
            "泵及真空设备",
            "2026-12-31",
            "是",
            "需具备独立法人资格；提供近3年同类业绩",
            120000,
        ]
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row, idx: int) -> str:
    v = row[idx] if idx < len(row) else None
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(raw: str) -> date | None:
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw[:10], fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(raw: str) -> float | None:
    if not raw:
        return None
    try:
        return float(raw.replace(",", ""))
    except ValueError:
        return None


def _parse_bool(raw: str) -> bool | None:
    s = (raw or "").strip().lower()
    if not s:
        return None
    if s in ("1", "true", "yes", "y", "是", "专门面向中小企业", "面向中小企业"):
        return True
    if s in ("0", "false", "no", "n", "否", "不面向", "非专门面向中小企业"):
        return False
    return None


def preview_excel(content: bytes) -> PlatformTenderExcelPreviewOut:
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows_out: list[PlatformTenderExcelPreviewRow] = []
    valid = 0
    errors = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = list(row)
        data = {
            "buyer_name": _cell(cells, 0),
            "industry": _cell(cells, 1) or None,
            "region": _cell(cells, 2) or None,
            "product_name": _cell(cells, 3) or None,
            "quantity": _cell(cells, 4) or None,
            "budget_min": _parse_float(_cell(cells, 5)),
            "budget_max": _parse_float(_cell(cells, 6)),
            "deadline": _parse_date(_cell(cells, 7)),
            "contact_name": _cell(cells, 8) or None,
            "contact_phone": _cell(cells, 9) or None,
            "source_url": _cell(cells, 10) or None,
            "summary": _cell(cells, 11) or None,
            "project_no": _cell(cells, 12) or None,
            "published_at": _parse_date(_cell(cells, 13)),
            "procurement_method": _cell(cells, 14) or None,
            "agent_name": _cell(cells, 15) or None,
            "buyer_address": _cell(cells, 16) or None,
            "category": _cell(cells, 17) or None,
            "bid_open_date": _parse_date(_cell(cells, 18)),
            "sme_preference": _parse_bool(_cell(cells, 19)),
            "qualification_summary": _cell(cells, 20) or None,
            "max_price_limit": _parse_float(_cell(cells, 21)),
        }
        errs: list[str] = []
        if not data["buyer_name"]:
            errs.append("采购方必填")
        if not data["source_url"]:
            errs.append("原文链接必填")
        if data["deadline"] is None and _cell(cells, 7):
            errs.append("截止日期格式无效")
        if data["published_at"] is None and _cell(cells, 13):
            errs.append("发布时间格式无效")
        if data["bid_open_date"] is None and _cell(cells, 18):
            errs.append("开标日期格式无效")
        if errs:
            errors += 1
        else:
            valid += 1
        dump = dict(data)
        for dk in ("deadline", "published_at", "bid_open_date"):
            if isinstance(dump.get(dk), date):
                dump[dk] = dump[dk].isoformat()
        rows_out.append(PlatformTenderExcelPreviewRow(row_num=i, data=dump, errors=errs))
    return PlatformTenderExcelPreviewOut(rows=rows_out, valid_count=valid, error_count=errors)


def confirm_excel(db: Session, admin: User, content: bytes) -> PlatformTenderExcelConfirmOut:
    preview = preview_excel(content)
    created_ids: list[UUID] = []
    skipped = 0
    for row in preview.rows:
        if row.errors:
            skipped += 1
            continue
        d = row.data

        def _d(key: str) -> date | None:
            v = d.get(key)
            if isinstance(v, str) and v:
                return date.fromisoformat(v)
            return None

        lead = create_lead(
            db,
            admin,
            PlatformTenderLeadCreate(
                buyer_name=d["buyer_name"],
                industry=d.get("industry"),
                region=d.get("region"),
                product_name=d.get("product_name"),
                quantity=d.get("quantity"),
                budget_min=d.get("budget_min"),
                budget_max=d.get("budget_max"),
                deadline=_d("deadline"),
                contact_name=d.get("contact_name"),
                contact_phone=d.get("contact_phone"),
                source_url=d.get("source_url"),
                summary=d.get("summary"),
                project_no=d.get("project_no"),
                published_at=_d("published_at"),
                procurement_method=d.get("procurement_method"),
                agent_name=d.get("agent_name"),
                buyer_address=d.get("buyer_address"),
                category=d.get("category"),
                bid_open_date=_d("bid_open_date"),
                sme_preference=d.get("sme_preference"),
                qualification_summary=d.get("qualification_summary"),
                max_price_limit=d.get("max_price_limit"),
                source_channel="excel",
                status="draft",
                has_source_document=True,
            ),
        )
        created_ids.append(lead.id)
    return PlatformTenderExcelConfirmOut(created=len(created_ids), skipped=skipped, ids=created_ids)
