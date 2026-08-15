import openpyxl, sys
from pathlib import Path
out = str(Path(__file__).resolve().parent / "plat_cases.txt")
p = r"C:\Users\admin\Desktop\临时\AI内容营销系统-测试用例.xlsx"
wb = openpyxl.load_workbook(p, data_only=True)
lines = []
for name in ["平台管理后台", "平台管理"]:
    if name not in wb.sheetnames:
        continue
    ws = wb[name]
    lines.append("SHEET: " + name + " rows=" + str(ws.max_row-1))
    for r in range(2, ws.max_row+1):
        cid = ws.cell(r, 2).value
        title = ws.cell(r, 6).value
        typ = ws.cell(r, 11).value
        step = ws.cell(r, 8).value
        exp = ws.cell(r, 9).value
        pri = ws.cell(r, 7).value
        if cid:
            lines.append(f"[{cid}] ({typ}/{pri}) {title}")
            if step: lines.append("  STEP: " + str(step).replace("\n", " | "))
            if exp: lines.append("  EXP: " + str(exp).replace("\n", " | "))
with open(out, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print("WROTE", out, "lines=", len(lines))
