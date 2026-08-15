#!/usr/bin/env python3
"""更新内容获客平台功能实施清单 Excel，对齐 AI 内容工厂 + 交易闭环 + CRM 新方案。"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path(r"C:\Users\admin\Desktop\学习\方案\内容获客平台_功能实施清单.xlsx")
OUT_DIR = ROOT / "docs/01-PRD/21-内容获客商城-phase1"
OUT_XLSX = OUT_DIR / "内容获客平台_功能实施清单.xlsx"
OUT_MD = OUT_DIR / "内容获客平台_功能实施清单.md"

HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1677FF")
BODY_FONT = Font(name="微软雅黑", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

AI_PHASE_ROWS = [
    ["A0-01", "多平台文案/笔记/视频脚本", "AI内容工厂", "营销创作顾问 + 活动关联", "可生成、可审核、可关联 CRM 活动", "Phase 0", "已交付"],
    ["A0-02", "知识库与品牌语气", "AI内容工厂", "KnowledgeDocument + 人格锁定", "检索命中；输出语气一致", "Phase 0", "已交付"],
    ["A0-03", "文案合规自检", "AI内容工厂", "合规引擎 + 改稿建议", "敏感词拦截；可一键改稿", "Phase 0", "已交付"],
    ["A1-01", "创作成果复用导出", "AI×商城", "文案/脚本导出包", "商家可粘贴到商品详情；非 Mx 硬验收", "Phase 1 可选", ""],
    ["A1-02", "平台运营端", "交易", "P01–P11 含 P10/P11", "平台可管商家/审核/渠道/订阅", "Phase 1", ""],
    ["A2-01", "智能体×商城打通", "AI×商城", "课纲/详情/卖点→商品草稿", "带商品 ID 创作→上架草稿", "Phase 2 · v1.7", ""],
    ["A3-01", "商用配图与图文 ZIP", "AI内容工厂", "图库 API + license 元数据", "未授权图不可导出", "Phase 3 · v1.8", ""],
    ["A3-02", "素材选源闸/成片闸", "合规", "商用白名单 + 授权追溯", "导出包附带授权清单", "Phase 3 · v1.8", ""],
    ["A4-01", "短视频半自动成片", "AI内容工厂", "TTS + 商用 BGM + 字幕", "人审确认后导出", "Phase 4 · v1.9", ""],
    ["A4-02", "素材发布闸", "合规", "文案+素材双人审", "不承诺全自动公域发布", "Phase 4 · v1.9", ""],
    ["A5-01", "A/B 与效果回流", "AI内容工厂", "渠道归因 + 发布数据回流", "可对比内容效果", "Phase 5 · v2.0+", ""],
    ["A5-02", "Multi-Agent 编排", "AI内容工厂", "选题→创作→合规→导出", "与 CRM/商城数据联动", "Phase 5 · v2.0+", ""],
]

PHASE1_ROWS = [
    ["P1-01", "商品类型模型", "交易", "shop_products.type = course|digital|service", "可上架三类；课非唯一 · M4", "Phase 1", ""],
    ["P1-02", "专栏/单课", "内容", "shop_columns/lessons", "tenant 隔离；CRUD · M4", "Phase 1", ""],
    ["P1-03", "数字权益商品", "内容", "资料包 SKU + 下载履约", "支付后可领取 · M4", "Phase 1", ""],
    ["P1-04", "服务预约/次数卡", "履约", "预约时段 + 核销/扣次", "核销后状态正确；退款关权益 · M6", "Phase 1", ""],
    ["P1-05", "买家/enrollment", "用户", "buyer/enrollment + crm_activities 桥接", "禁止硬映射 Contact", "Phase 1", ""],
    ["P1-06", "统一权益状态机", "交易", "shop_entitlements", "开通/核销/到期/退款撤销 · M3/M5", "Phase 1", ""],
    ["P1-07", "订单与支付", "交易", "shop_orders + 微信支付 + 幂等", "M3 硬验收", "Phase 1", ""],
    ["P1-08", "退款撤销权益", "交易", "退款回调 → 关权益", "退款后立即不可履约 · M5", "Phase 1", ""],
    ["P1-09", "小程序履约端", "私域", "学课 / 核销 / 已购入口", "签约客户可运营", "Phase 1", ""],
    ["P1-10", "视频/音频播放", "内容", "倍速/续播/试看", "课类 SKU 可用", "Phase 1", ""],
    ["P1-11", "抖音公域 Mx · 链路 ①", "公域", "抖店 Webhook；路径 A/B", "Mx 抖店付→领权→履约 · M7", "Phase 1", ""],
    ["P1-11b", "抖音公域 Mx · 链路 ②", "公域", "课程库+小程序交易 API", "与 ① 二选一先通 · M8 可选", "Phase 1", ""],
    ["P1-12", "领权短信", "私域", "手机号兑换/领权链路", "公域订单可进自有端 · M7", "Phase 1", ""],
    ["P1-13", "Web 管理端", "管理", "A01–A22 商品/订单/买家/权益", "商家可运营", "Phase 1", ""],
    ["P1-14", "交易看板", "数据", "A01 · §8.15.1 看板 API", "数据准确；下钻 A09", "Phase 1", ""],
    ["P1-15", "可选营客台桥接", "经营", "shop 事件进时间线", "非强制；可演示", "Phase 1", ""],
    ["P1-16", "IP 获客演示包", "获客", "CTA → 挂载 → 下单脚本", "可独立演示；index#p1-16-demo · PRD §9.3", "Phase 1", ""],
    ["P1-17", "C 端发票申请", "交易", "shop_invoice_requests", "买家可提交；商家可见状态 · M6", "Phase 1", ""],
    ["P1-18", "买家订单中心", "私域", "M11/M12 订单详情/退款/开票", "与已购分离 · M5", "Phase 1", ""],
    ["P1-19", "商品上架合规闸", "合规", "机审+人审；pending_review", "驳回不可 on_sale · M4", "Phase 1", ""],
    ["P1-20", "公域挂载合规闸", "公域", "A14 映射前校验 F7", "未过审不可同步 · M7", "Phase 1", ""],
    ["P1-21", "商品合规审计日志", "合规", "审核人/时间/驳回原因留痕", "P09 驳回原因留痕 · M4", "Phase 1", ""],
]

OURS_NEW = [
    [
        "AI 内容工厂（生产层）",
        "仅有「再创作」，缺从 0 到 1 的生产闭环",
        "公域获客靠内容，不能只卖「店」；要先帮商家写得出来",
        "营销顾问 v0.6～v1.6 已交付；v1.7 起与商城打通",
        "Phase 0 ✅；Phase 1–5 分期",
    ],
    [
        "三位一体闭环",
        "课店与履约强；生产与经营弱",
        "生产（AI）→ 成交（商城）→ 经营（CRM）一条链",
        "智营 CRM 已有；商城 Phase1 立项",
        "Phase 0 ✅ + Phase 1 主攻",
    ],
    [
        "素材合规三道闸",
        "版权防护偏水印；无系统授权追溯",
        "公域配图/BGM 侵权易限流封号",
        "选源闸·成片闸·发布闸 + license 元数据",
        "Phase 3/4 建设",
    ],
    [
        "商品上架+公域挂载合规",
        "有内容审核，但与挂店/课程库提审耦合弱",
        "挂课卖虚拟服务是教培强监管场景；不过审不能卖、更不能挂公域",
        "上架闸机审+人审；公域挂载闸；审核版本锁定",
        "Phase 1 · P0",
    ],
]

NOT_DO_NEW = [
    ["全自动公域发布", "部分平台支持", "各平台政策差异大；合规与人审不可省"],
    ["无授权 AI 图/流行曲商用", "—", "侵权与限流风险；仅商用白名单"],
    ["AI 数字人直播", "有探索", "投入产出比低；非 Phase 1–3 重点"],
]

README_LINES = [
    "内容获客平台 · 功能清单（导出自小鹅通对标 + AI 内容工厂新方案）",
    "",
    "产品定位：AI 内容工厂 + 交易闭环 + CRM 持续经营",
    "平台三层：",
    "  · AI 内容工厂（生产层）— 文案/笔记/脚本；后续图文/短视频",
    "  · 内容获客平台（交易层）— 商品/订单/支付/权益；四平台挂载",
    "  · 智营 CRM（运营层）— 活动/线索/客户；成交后跟进",
    "",
    "分期说明：",
    "  · Phase 0（AI）✅ 已交付 — 营销创作顾问",
    "  · Phase 1（商城）— 虚拟/服务交易闭环 + 抖音 Mx",
    "  · Phase 2–5（AI 增强）— 商城打通 → 配图 → 短视频 → 效果闭环",
    "  · 商城 Phase 2/3 — 装修/H5/实物/四平台（见各实施清单表）",
    "",
    "硬门槛：M3 私域支付+权益；M2 含 A20 自申；Mx 抖音公域首单（M7/M8 · 链路 ① 或 ②）",
    "验收对照：PRD §九 · §1.1 · §8.13（verify_shop_m0.py 已覆盖 M0）",
    "路径：A 官方店 / B 客户自有店（与公域链路不同维度）",
    "C 端买家：shop_buyers；≠ B2B Contact",
    "",
    "工作表：",
    "1) AI内容工厂分期 — 智能体 Phase 0～5",
    "2) Phase1实施清单 — 虚拟/服务 + 抖音首单",
    "3) Phase2/3实施清单 — 扩公域 + 开店 + 实物/门店",
    "4) 我方独有能力 — 含 AI 内容工厂差异化",
    "5) 明确不做",
    "6) 总对标矩阵 / 对标明细汇总",
    "7) 实施清单合并 — 全量一张表",
    "",
    f"更新日期：{date.today().isoformat()}",
    "关联：docs/00-总览/内容获客平台方案摘要.md",
    "       docs/01-PRD/21-内容获客商城-phase1/PRD-内容获客商城-phase1.md §九",
]


def style_header_row(ws, row: int = 1) -> None:
    for cell in ws[row]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = WRAP


def style_body(ws, start_row: int = 2) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            if cell.value is not None:
                cell.font = BODY_FONT
                cell.alignment = WRAP


def append_rows(ws, rows: list[list], start: int | None = None) -> None:
    if start is None:
        start = ws.max_row + 1
    for i, data in enumerate(rows):
        for j, val in enumerate(data, 1):
            ws.cell(start + i, j, val)


def replace_readme(ws) -> None:
    ws.delete_rows(1, ws.max_row)
    for i, line in enumerate(README_LINES, 1):
        c = ws.cell(i, 1, line)
        c.font = Font(name="微软雅黑", size=11, bold=(i == 1))
        c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 100


def upsert_sheet(wb, name: str, headers: list[str], rows: list[list]) -> None:
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    style_header_row(ws)
    style_body(ws)
    for col, w in zip("ABCDEFGH", [10, 22, 12, 28, 36, 14, 12, 10]):
        ws.column_dimensions[col].width = w


def dedupe_ours_sheet(ws) -> None:
    seen: set[str] = set()
    to_delete: list[int] = []
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if not raw:
            continue
        key = str(raw).split("|")[0].strip()
        if key in seen:
            to_delete.append(r)
        else:
            seen.add(key)
    for r in reversed(to_delete):
        ws.delete_rows(r)


def update_ours(ws) -> None:
    # 更新 IP / 多品类 行
    for r in range(2, ws.max_row + 1):
        cap = ws.cell(r, 1).value
        if not cap:
            continue
        if "IP 内容矩阵" in str(cap):
            ws.cell(r, 4).value = "AI 内容工厂 + 挂载 CTA + 订单回流看板"
            ws.cell(r, 5).value = "Phase 0 生产已交付；Phase 1 演示包 + 交易闭环"
        if "多品类商品类型" in str(cap):
            ws.cell(r, 1).value = "多品类商品类型"
            ws.cell(r, 4).value = "shop_products.type：course、digital、service（后期 +physical / +store）"
        if "品类演进" in str(cap):
            ws.cell(r, 4).value = "shop_products.type 扩展 + 物流/到店核销插件"
    existing_caps = {
        str(ws.cell(r, 1).value).split("|")[0].strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value
    }
    to_add = [row for row in OURS_NEW if row[0] not in existing_caps]
    if to_add:
        append_rows(ws, to_add)
    dedupe_ours_sheet(ws)


def update_not_do(ws) -> None:
    append_rows(ws, NOT_DO_NEW)


def rebuild_merge(wb) -> None:
    headers = ["序号", "功能项", "模块", "交付物", "验收/阶段说明", "阶段", "类型", "状态"]
    rows: list[list] = []
    for sn, typ in [
        ("AI内容工厂分期", "AI智能体"),
        ("Phase1实施清单", "实施项"),
        ("Phase2实施清单", "实施项"),
        ("Phase3实施清单", "实施项"),
    ]:
        ws = wb[sn]
        for r in range(2, ws.max_row + 1):
            seq = ws.cell(r, 1).value
            if not seq:
                continue
            rows.append([
                seq,
                ws.cell(r, 2).value,
                ws.cell(r, 3).value,
                ws.cell(r, 4).value,
                ws.cell(r, 5).value,
                ws.cell(r, 6).value,
                typ,
                ws.cell(r, 7).value if ws.max_column >= 7 else "",
            ])
    ours = wb["我方独有能力"]
    for r in range(2, ours.max_row + 1):
        cap = ours.cell(r, 1).value
        if not cap:
            continue
        rows.append([
            "OURS",
            cap,
            "独有",
            ours.cell(r, 4).value,
            ours.cell(r, 5).value,
            ours.cell(r, 6).value or "见阶段列",
            "小鹅通没有·我方需要",
            "",
        ])
    if "实施清单合并" in wb.sheetnames:
        del wb["实施清单合并"]
    upsert_sheet(wb, "实施清单合并", headers, rows)


def update_matrix(ws) -> None:
    # AI能力：已有 3 -> 5（Phase0 三项 + 销售跟进 + 内容生成已有）
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "AI能力":
            ws.cell(r, 2).value = 5  # 已有
            ws.cell(r, 3).value = 1  # Phase1 复用导出
            ws.cell(r, 4).value = 3  # Phase2 打通+助理+诊断
            ws.cell(r, 5).value = 4  # Phase3-5 配图/成片/闭环
            break
    # 重算合计
    totals = [0] * 5
    for r in range(2, ws.max_row):
        mod = ws.cell(r, 1).value
        if mod == "合计":
            for c in range(2, 7):
                ws.cell(r, c).value = totals[c - 2]
            break
        for c in range(2, 7):
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                totals[c - 2] += int(v)


def export_markdown(wb) -> None:
    lines = [
        "# 内容获客平台 · 功能实施清单",
        "",
        "> Excel 版：[内容获客平台_功能实施清单.xlsx](./内容获客平台_功能实施清单.xlsx) · "
        "方案摘要：[../../00-总览/内容获客平台方案摘要.md](../../00-总览/内容获客平台方案摘要.md)  ",
        "> **验收对照**：[PRD §九 验收与测试](./PRD-内容获客商城-phase1.md#九验收与测试) · "
        "[§1.1 范围总表](./PRD-内容获客商城-phase1.md#11-phase-1-范围总表a1--p1--) · "
        "[§8.13 实现批次](./PRD-内容获客商城-phase1.md#813-实现批次对照)",
        "",
        f"更新：{date.today().isoformat()}",
        "",
        "## AI 内容工厂分期（Phase 0～5）",
        "",
        "| 序号 | 功能项 | 模块 | 交付物 | 验收标准 | 阶段 | 状态 |",
        "|------|--------|------|--------|----------|------|------|",
    ]
    ws = wb["AI内容工厂分期"]
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        cols = [ws.cell(r, c).value or "" for c in range(1, 8)]
        lines.append("| " + " | ".join(str(c) for c in cols) + " |")
    lines += ["", "## Phase 1 实施清单（交易层首期）", "", "| 序号 | 功能项 | 模块 | 交付物 | 验收标准 | 阶段 | 状态 |", "|------|--------|------|--------|----------|------|------|"]
    ws = wb["Phase1实施清单"]
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 1).value:
            continue
        cols = [ws.cell(r, c).value or "" for c in range(1, 8)]
        lines.append("| " + " | ".join(str(c) for c in cols) + " |")
    lines += [
        "",
        "## 我方独有能力（摘要）",
        "",
        "| 能力 | 我方落点 | 阶段 |",
        "|------|----------|------|",
    ]
    for r in range(2, wb["我方独有能力"].max_row + 1):
        raw_cap = wb["我方独有能力"].cell(r, 1).value
        if not raw_cap or str(raw_cap).startswith("能力"):
            continue
        cap = str(raw_cap).split("|")[0].strip() if "|" in str(raw_cap) else str(raw_cap).strip()
        stage = wb["我方独有能力"].cell(r, 5).value or ""
        landing = wb["我方独有能力"].cell(r, 4).value or ""
        lines.append(f"| {cap} | {landing} | {stage} |")
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    src = DESKTOP if DESKTOP.exists() else OUT_XLSX
    if not src.exists():
        raise SystemExit(f"找不到 Excel：{DESKTOP} 或 {OUT_XLSX}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        shutil.copy2(src, OUT_XLSX)
        load_path = OUT_XLSX
    except PermissionError:
        load_path = src
        print(f"目标 Excel 被占用，直接更新：{load_path}")

    wb = openpyxl.load_workbook(load_path)
    replace_readme(wb["说明"])

    upsert_sheet(
        wb,
        "AI内容工厂分期",
        ["序号", "功能项", "模块", "交付物", "验收标准", "阶段", "状态"],
        AI_PHASE_ROWS,
    )

    update_ours(wb["我方独有能力"])
    update_not_do(wb["明确不做"])

    upsert_sheet(
        wb,
        "Phase1实施清单",
        ["序号", "功能项", "模块", "交付物", "验收标准", "阶段", "状态"],
        PHASE1_ROWS,
    )
    # 若从旧模板加载，upsert 已全量替换；双保险再扫一遍 product.type
    if "Phase1实施清单" in wb.sheetnames:
        for r in range(2, wb["Phase1实施清单"].max_row + 1):
            cell = wb["Phase1实施清单"].cell(r, 4)
            if cell.value and "product.type" in str(cell.value):
                cell.value = str(cell.value).replace("product.type", "shop_products.type")

    update_matrix(wb["总对标矩阵"])
    rebuild_merge(wb)

    # 各表统一字体
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn != "说明":
            if ws.max_row >= 1 and ws.cell(1, 1).value:
                style_header_row(ws)
            style_body(ws, 2 if sn != "说明" else 1)

    wb.save(OUT_XLSX)
    export_markdown(wb)
    print(f"已更新：{OUT_XLSX}")
    print(f"已生成：{OUT_MD}")


if __name__ == "__main__":
    main()
